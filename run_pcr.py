"""
PCR — Performance Calculator and Ranking — runner script.

Loads a master Excel file, scores every 2025+ run with PCR, and writes
an Excel report. PCR is a standalone analytics module; this runner is
project-independent (does not touch PA1 reports).

By default uses the **Shared / Teams-synced master** (Friday-QC'd file)
since ROP / drill / motor entries have been cleaned by the team and are
more reliable than the Local pre-QC copies.

Usage:
  python run_pcr.py                                 # interactive: pick source + confirm
  python run_pcr.py --source shared --week 26-W17 --yes
  python run_pcr.py --source local --week 26-W17    # explicitly use pre-QC local copy
  python run_pcr.py --file path/to/master.xlsx
  python run_pcr.py --output my_rankings.xlsx

Methodology: docs/pcr.md
Config:      config/pcr.yaml
Module:      src/pcr.py
"""

import argparse
import os
import sys
import time
from datetime import datetime

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from src.data_loader import (
    _copy_to_temp, find_master_file_interactive,
    find_master_file_local, load_and_clean, load_config,
)
from src.pcr import rank_runs


# =====================================================================
# Interactive helpers (source picker + file confirm)
# =====================================================================

# PCR uses the Shared / Teams-synced file by default — those are the
# Friday-QC'd masters where ROP / drill / motor data has been cleaned. The
# Local OneDrive copies are pre-QC and may contain raw entry errors that
# would skew the rankings. To override (e.g. for an ad-hoc analysis on a
# specific file), pass --source local or --file.
def _choose_source_interactive():
    print("\n  Select master file source (default: shared/QC'd):")
    print("    [1] Shared -- Teams sync, Friday-QC'd  (RECOMMENDED — default)")
    print("    [2] Local  -- your OneDrive copy (pre-QC, may contain raw errors)")
    print("    [3] Browse -- pick any file from a list (all paths)")
    choice = input("  > ").strip().lower()
    if choice in ("", "1", "shared"):
        return "shared"
    if choice in ("2", "local"):
        return "local"
    if choice in ("3", "browse"):
        return "browse"
    print(f"  Unrecognized choice '{choice}', defaulting to shared.")
    return "shared"


def _confirm_file(original, display_path):
    print("\n  ============================================================")
    print(f"  Detected file: {original}")
    print(f"  Path:          {display_path}")
    if os.path.exists(display_path):
        mtime = datetime.fromtimestamp(os.path.getmtime(display_path)).strftime("%Y-%m-%d %H:%M")
        size_mb = os.path.getsize(display_path) / 1024 / 1024
        print(f"  Modified:      {mtime}  ({size_mb:.1f} MB)")
    print("  ============================================================")
    return input("  Proceed with this file? [Y/n]: ").strip().lower() in ("", "y", "yes")


def _resolve_master_file(source, config):
    if source == "local":
        local = find_master_file_local(config, scope="local")
        return _copy_to_temp(local), os.path.basename(local), local
    if source == "shared":
        local = find_master_file_local(config, scope="shared")
        return _copy_to_temp(local), os.path.basename(local), local
    if source == "browse":
        local, original = find_master_file_interactive(config, scope="all")
        return _copy_to_temp(local), original, local
    raise ValueError(f"Unknown source: {source}")


# =====================================================================
# Excel writer — three sheets: This Week / All Scored / Weekly Aggregate
# =====================================================================

# Columns shown in This Week and All Scored sheets (in order).
DETAIL_COLS = [
    "DATE_OUT", "Week #", "OPERATOR", "WELL", "JOB_NUM",
    "HOLE_SIZE", "Phase_CALC", "FOOTAGE_BUCKET",
    "COUNTY", "FORMATION", "BASIN",
    "MOTOR_TYPE2", "MOTOR_MODEL", "SERIES 20",
    "TOTAL_DRILL", "AVG_ROP", "ROTATE_ROP", "SLIDE_ROP",
    "pcr_score", "pcr_rank", "pcr_band",
    "pcr_peer_level", "pcr_peer_n",
    "pcr_total_drill_pct", "pcr_avg_rop_pct",
    "pcr_rotate_rop_pct", "pcr_slide_rop_pct",
    "peer_avg_pcr_365d", "delta_vs_365d",
]

# Per-column widths for nicer Excel rendering.
COL_WIDTH = {
    "DATE_OUT": 11, "Week #": 8, "OPERATOR": 28, "WELL": 30, "JOB_NUM": 9,
    "HOLE_SIZE": 9, "Phase_CALC": 11, "FOOTAGE_BUCKET": 11,
    "COUNTY": 16, "FORMATION": 14, "BASIN": 14,
    "MOTOR_TYPE2": 12, "MOTOR_MODEL": 12, "SERIES 20": 8,
    "TOTAL_DRILL": 11, "AVG_ROP": 10, "ROTATE_ROP": 11, "SLIDE_ROP": 10,
    "pcr_score": 10, "pcr_rank": 8, "pcr_band": 10,
    "pcr_peer_level": 11, "pcr_peer_n": 9,
    "pcr_total_drill_pct": 13, "pcr_avg_rop_pct": 12,
    "pcr_rotate_rop_pct": 13, "pcr_slide_rop_pct": 12,
    "peer_avg_pcr_365d": 14, "delta_vs_365d": 12,
}

BAND_FILL = {
    "Excellent": "63BE7B",
    "Good":      "BFE0B0",
    "Average":   "FFEB84",
    "Poor":      "F8696B",
}


def _format_cell(cell, col):
    if col == "pcr_score":
        cell.number_format = "0.0"
    elif col == "pcr_rank":
        cell.number_format = "0"
    elif col == "pcr_peer_n":
        cell.number_format = "0"
    elif col == "TOTAL_DRILL":
        cell.number_format = "#,##0"
    elif col.endswith("_pct"):
        cell.number_format = "0.0"
    elif col == "DATE_OUT":
        cell.number_format = "yyyy-mm-dd"
    elif col in ("AVG_ROP", "ROTATE_ROP", "SLIDE_ROP",
                 "peer_avg_pcr_365d", "delta_vs_365d"):
        cell.number_format = "0.00"


def _write_detail_sheet(ws, df, title=None):
    bold_white = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cols = [c for c in DETAIL_COLS if c in df.columns]

    start_row = 1
    if title:
        c = ws.cell(row=1, column=1, value=title)
        c.font = Font(bold=True, size=12)
        start_row = 3

    # Headers
    for ci, col in enumerate(cols, start=1):
        h = ws.cell(row=start_row, column=ci, value=col)
        h.font = bold_white
        h.fill = header_fill
        h.alignment = center_wrap
    ws.row_dimensions[start_row].height = 28

    # Data
    for ri, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for ci, col in enumerate(cols, start=1):
            val = row.get(col)
            cell = ws.cell(row=ri, column=ci, value=None if pd.isna(val) else val)
            _format_cell(cell, col)
            # Color the band cell by band value
            if col == "pcr_band" and isinstance(val, str) and val in BAND_FILL:
                cell.fill = PatternFill("solid", fgColor=BAND_FILL[val])
                cell.alignment = Alignment(horizontal="center")

    # Column widths
    for ci, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTH.get(col, 12)

    n_data = len(df)
    if n_data == 0:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate
        return

    # Conditional formatting: red→yellow→green on pcr_score, white-centered on delta
    if "pcr_score" in cols:
        col_idx = cols.index("pcr_score") + 1
        letter = get_column_letter(col_idx)
        rule = ColorScaleRule(
            start_type="num", start_value=0, start_color="F8696B",
            mid_type="num", mid_value=50, mid_color="FFEB84",
            end_type="num", end_value=100, end_color="63BE7B",
        )
        ws.conditional_formatting.add(f"{letter}{start_row+1}:{letter}{start_row+n_data}", rule)

    if "delta_vs_365d" in cols:
        col_idx = cols.index("delta_vs_365d") + 1
        letter = get_column_letter(col_idx)
        rule = ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="max", end_color="63BE7B",
        )
        ws.conditional_formatting.add(f"{letter}{start_row+1}:{letter}{start_row+n_data}", rule)

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate


def _write_aggregate_sheet(ws, scored):
    bold_white = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    center = Alignment(horizontal="center", vertical="center")

    weekly = (
        scored.groupby("Week #")
        .agg(
            n_runs=("pcr_score", "count"),
            avg_pcr=("pcr_score", "mean"),
            median_pcr=("pcr_score", "median"),
            min_pcr=("pcr_score", "min"),
            max_pcr=("pcr_score", "max"),
            avg_peer_365d=("peer_avg_pcr_365d", "mean"),
            avg_delta=("delta_vs_365d", "mean"),
        )
        .reset_index()
        .sort_values("Week #", ascending=False)
    )

    cols = list(weekly.columns)
    for ci, col in enumerate(cols, start=1):
        h = ws.cell(row=1, column=ci, value=col)
        h.font = bold_white
        h.fill = header_fill
        h.alignment = center
    ws.row_dimensions[1].height = 25

    for ri, (_, row) in enumerate(weekly.iterrows(), start=2):
        for ci, col in enumerate(cols, start=1):
            val = row.get(col)
            cell = ws.cell(row=ri, column=ci, value=None if pd.isna(val) else val)
            if col == "n_runs":
                cell.number_format = "0"
            elif col != "Week #":
                cell.number_format = "0.00"

    widths = {"Week #": 9, "n_runs": 8}
    for ci, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 14)
    ws.freeze_panes = "A2"

    # Color avg_pcr column with the same rule
    if "avg_pcr" in cols and len(weekly) > 0:
        col_idx = cols.index("avg_pcr") + 1
        letter = get_column_letter(col_idx)
        rule = ColorScaleRule(
            start_type="num", start_value=30, start_color="F8696B",
            mid_type="num", mid_value=50, mid_color="FFEB84",
            end_type="num", end_value=70, end_color="63BE7B",
        )
        ws.conditional_formatting.add(f"{letter}2:{letter}{1+len(weekly)}", rule)


def write_pcr_excel(ranked, week, output_path):
    scored = ranked[ranked["pcr_score"].notna()].copy()
    scored = scored.sort_values("pcr_score", ascending=False)

    wb = Workbook()

    # Sheet 1 — This Week
    ws1 = wb.active
    ws1.title = "This Week"
    week_df = scored[scored["Week #"] == week] if week else scored.head(0)
    title = f"PCR Rankings — Week {week}" if week else "PCR Rankings"
    _write_detail_sheet(ws1, week_df, title=title)

    # Sheet 2 — All Scored
    ws2 = wb.create_sheet("All Scored")
    _write_detail_sheet(ws2, scored)

    # Sheet 3 — Weekly Aggregate
    ws3 = wb.create_sheet("Weekly Aggregate")
    _write_aggregate_sheet(ws3, scored)

    wb.save(output_path)


# =====================================================================
# Main
# =====================================================================

def main():
    parser = argparse.ArgumentParser(description="PCR — Performance Calculator and Ranking")
    parser.add_argument("--source", choices=["local", "shared", "browse"], default=None)
    parser.add_argument("--week", default=None, help="Week ID (e.g. 26-W17) — used for 'This Week' sheet")
    parser.add_argument("--file", default=None)
    parser.add_argument("--output", default=None)
    parser.add_argument("--yes", "-y", action="store_true", help="Skip file confirmation")
    args = parser.parse_args()

    print("[1/5] Loading configs...")
    cfg_data = load_config()
    pcr_yaml = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config", "pcr.yaml")
    with open(pcr_yaml, "r") as f:
        cfg_pcr = yaml.safe_load(f)

    print("[2/5] Locating master file...")
    if args.file:
        if not os.path.exists(args.file):
            print(f"  ERROR: File not found: {args.file}")
            sys.exit(1)
        read_path = _copy_to_temp(args.file)
        original = os.path.basename(args.file)
        display_path = args.file
    else:
        source = args.source or _choose_source_interactive()
        print(f"  Source: {source}")
        read_path, original, display_path = _resolve_master_file(source, cfg_data)

    if not args.yes and not _confirm_file(original, display_path):
        print("  Aborted.")
        sys.exit(0)

    print(f"  Master: {original}")

    print("[3/5] Loading data...")
    df = load_and_clean(read_path, cfg_data)
    print(f"  Loaded {len(df):,} rows.")

    # Determine "This Week" filter
    week = args.week
    if not week and "Week #" in df.columns:
        weeks = [w for w in df["Week #"].dropna().unique() if isinstance(w, str) and "-W" in w]
        if weeks:
            week = max(weeks)
            print(f"  Auto-detected latest week: {week}")

    print("[4/5] Scoring runs with PCR...")
    t0 = time.time()
    ranked = rank_runs(df, cfg_pcr)
    print(f"  Done in {time.time() - t0:.1f}s")
    n_scored = ranked["pcr_score"].notna().sum()
    print(f"  Scored: {n_scored:,} of {len(ranked):,} runs")

    print("[5/5] Writing Excel...")
    output_path = args.output or f"PCR Ranking - Week {week}.xlsx"
    try:
        write_pcr_excel(ranked, week, output_path)
    except PermissionError:
        print(f"\n  ERROR: '{output_path}' is open. Close it and re-run.")
        sys.exit(1)
    print(f"  Saved: {output_path}")
    print(f"  Sheets: This Week (week {week}) | All Scored ({n_scored:,} runs) | Weekly Aggregate")


if __name__ == "__main__":
    main()
