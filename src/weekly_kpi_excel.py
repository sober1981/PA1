"""Excel writer for the Weekly KPI Summary table.

Layout:
  - Summary Table at the top (per-hole-size totals + Grand Total).
  - Detailed Table below: per-(MOTOR_TYPE2 / JOB_TYPE / SERIES 20) rows + Longest Run
    for each hole size, wrapped in a thick black border per hole size.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


SUMMARY_HEADERS = [
    "Week #", "Hole Size", "Total Runs", "Total Hrs",
    "Hrs Diff vs Prev Week", "% Of G total (Hrs)",
    "Total Drill", "Ftg vs Prev Week", "% Of G total (Drill)",
    "Total Incidents", "OP w/ More Runs",
]
SUMMARY_LAST_COL = 11  # Summary table spans cols A..K

DETAILED_HEADERS = [
    "Week #", "Hole Size", "Motor Type", "Job Type", "Series 20",
    "Total Runs", "Total Hrs", "% Of W total (Hrs)", "% Of G total (Hrs)",
    "Total Drill", "% Of W total (Drill)", "% Of G total (Drill)",
    "Drilling Hrs", "Avg ROP", "Avg Slide %", "Avg Run Length",
    "MY Avg", "Incident Count", "OP w/ More Runs",
]
DETAILED_LAST_COL = 19  # Detailed table spans cols A..S

# Longest Run Table — packed sequential layout (cols A..P)
LONGEST_RUN_HEADERS = [
    "Week #", "Hole Size", "Motor Type", "Job Type", "Series 20",
    "Total Hrs", "Total Drill", "Drilling Hrs", "Avg ROP", "Avg Slide %",
    "MY Avg", "Incident Count",
    "Operator", "Job Number", "Phase", "Bend",
]
# 1-based positions in the packed layout
LR_COL_WEEK         = 1
LR_COL_HOLE_SIZE    = 2
LR_COL_MOTOR_TYPE   = 3
LR_COL_JOB_TYPE     = 4
LR_COL_SERIES_20    = 5
LR_COL_TOTAL_HRS    = 6
LR_COL_TOTAL_DRILL  = 7
LR_COL_DRILLING_HRS = 8
LR_COL_AVG_ROP      = 9
LR_COL_AVG_SLIDE    = 10
LR_COL_MY_AVG       = 11
LR_COL_INCIDENT     = 12
LR_COL_OPERATOR     = 13
LR_COL_JOB_NUM      = 14
LR_COL_PHASE        = 15
LR_COL_BEND         = 16
LONGEST_LAST_COL    = 16

# Widths apply to all three tables (shared column letters).
# Col M (13) -> Operator in Longest Run / Drilling Hrs in Detailed.
# Col S (19) -> Detailed's "OP w/ More Runs" (22 wide).
COL_WIDTHS = [8, 11, 13, 13, 9, 11, 11, 13, 13, 12, 14, 14, 16, 12, 14, 10, 9, 9, 22]

MOTOR_FILL = {
    "TDI CONV":   "D5E8D4",
    "CAM RENTAL": "E7E6E6",
    "CAM DD":     "F4CCCC",
    "3RD PARTY":  "CFE2F3",
}

MOTOR_FILL_DARK = {
    "TDI CONV":   "548235",
    "CAM RENTAL": "595959",
    "CAM DD":     "9C0006",
    "3RD PARTY":  "1F4E78",
}

GREEN_BORDER = "00B050"
RED_BORDER = "C00000"
BLACK_BORDER = "000000"

GRAND_FILL = "D9D9D9"

# Column indices (1-based) — detailed table
COL_HOLE_SIZE = 2
COL_SERIES_20 = 5
COL_TOTAL_HRS = 7
COL_PCT_G_HRS = 9
COL_TOTAL_DRILL = 10
COL_PCT_G_DRILL = 12
COL_AVG_RUNLEN = 16
COL_COMMENT = 19

DATA_ROW_HEIGHT = 15
HEADER_ROW_HEIGHT = 30

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
COMMA_FMT = "#,##0"


def _week_num(week):
    if isinstance(week, str) and "-W" in week:
        try:
            return int(week.split("-W")[-1])
        except ValueError:
            return week
    return week


def _normalize_motor_key(motor_type):
    if not motor_type:
        return None
    mt = motor_type.upper().strip()
    return mt if mt in MOTOR_FILL else None


def _apply_border_span(ws, row, col_start, col_end, color, style="thick"):
    side = Side(border_style=style, color=color)
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        ex = cell.border
        cell.border = Border(
            left=(side if c == col_start else ex.left),
            right=(side if c == col_end else ex.right),
            top=side,
            bottom=side,
        )


def _apply_border_single(ws, row, col, color, style="thick"):
    side = Side(border_style=style, color=color)
    cell = ws.cell(row=row, column=col)
    cell.border = Border(left=side, right=side, top=side, bottom=side)


def _apply_block_border(ws, top_row, bottom_row, col_start, col_end, color, style="medium"):
    """Outline a 2D rectangle (top_row..bottom_row, col_start..col_end) with a colored border.
    Preserves any existing borders on inner edges."""
    side = Side(border_style=style, color=color)
    for r in range(top_row, bottom_row + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            ex = cell.border
            cell.border = Border(
                left=(side if c == col_start else ex.left),
                right=(side if c == col_end else ex.right),
                top=(side if r == top_row else ex.top),
                bottom=(side if r == bottom_row else ex.bottom),
            )


def _diff_font(value, bold=False):
    """Return a Font with red (neg) / green (pos) color, or None for zero/None."""
    if value is None or value == 0:
        return Font(bold=True) if bold else None
    color = "C00000" if value < 0 else "006100"
    return Font(bold=bold, color=color)


def _write_summary_table(ws, kpi_data, start_row, week_num):
    """Write the summary table starting at start_row. Returns the row immediately after."""
    bold = Font(bold=True)
    bold_white = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    grand_fill = PatternFill("solid", fgColor=GRAND_FILL)

    # Header row
    for c, h in enumerate(SUMMARY_HEADERS, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.font = bold_white
        cell.fill = header_fill
        cell.alignment = CENTER_WRAP
    ws.row_dimensions[start_row].height = HEADER_ROW_HEIGHT

    row = start_row + 1
    data_first_row = row
    totals = {"runs": 0, "total_hrs": 0.0, "hrs_diff": 0.0, "g_pct_hrs": 0.0,
              "total_drill": 0, "ftg_diff": 0, "g_pct_drill": 0.0, "incidents": 0}

    for block in kpi_data["blocks"]:
        ct = block["curr_total"]
        diff = block["diff"]
        ws.cell(row=row, column=1, value=week_num)
        ws.cell(row=row, column=2, value=block["hole_size"])
        ws.cell(row=row, column=3, value=ct["runs"])
        ws.cell(row=row, column=4, value=round(ct["total_hrs"], 2))
        c = ws.cell(row=row, column=5, value=round(diff["total_hrs"], 2))
        df_font = _diff_font(diff["total_hrs"])
        if df_font:
            c.font = df_font
        c = ws.cell(row=row, column=6, value=ct["g_pct_hrs"]); c.number_format = "0.00%"
        c = ws.cell(row=row, column=7, value=ct["total_drill"]); c.number_format = COMMA_FMT
        c = ws.cell(row=row, column=8, value=diff["total_drill"]); c.number_format = COMMA_FMT
        df_font = _diff_font(diff["total_drill"])
        if df_font:
            c.font = df_font
        c = ws.cell(row=row, column=9, value=ct["g_pct_drill"]); c.number_format = "0.00%"
        ws.cell(row=row, column=10, value=ct["incident_count"])
        ws.cell(row=row, column=11, value=ct.get("top_operator", ""))
        ws.row_dimensions[row].height = DATA_ROW_HEIGHT
        for c in range(1, len(SUMMARY_HEADERS) + 1):
            cell = ws.cell(row=row, column=c)
            if cell.alignment.horizontal != "center":
                cell.alignment = CENTER

        totals["runs"] += ct["runs"]
        totals["total_hrs"] += ct["total_hrs"]
        totals["hrs_diff"] += diff["total_hrs"]
        totals["g_pct_hrs"] += ct["g_pct_hrs"]
        totals["total_drill"] += ct["total_drill"]
        totals["ftg_diff"] += diff["total_drill"]
        totals["g_pct_drill"] += ct["g_pct_drill"]
        totals["incidents"] += ct["incident_count"]
        row += 1

    data_last_row = row - 1  # for color scale ranges

    # True grand-level diffs (current week vs prev week — across ALL hole sizes,
    # including any that exist only in one week)
    grand_hrs_diff = totals["total_hrs"] - kpi_data.get("grand_prev_total_hrs", 0.0)
    grand_ftg_diff = totals["total_drill"] - kpi_data.get("grand_prev_total_drill", 0)

    # Grand Total (no fill — clean white)
    for c in range(1, len(SUMMARY_HEADERS) + 1):
        ws.cell(row=row, column=c).alignment = CENTER
    ws.cell(row=row, column=2, value="Grand Total").font = bold
    ws.cell(row=row, column=3, value=totals["runs"]).font = bold
    ws.cell(row=row, column=4, value=round(totals["total_hrs"], 2)).font = bold
    c = ws.cell(row=row, column=5, value=round(grand_hrs_diff, 2))
    df_font = _diff_font(grand_hrs_diff, bold=True)
    c.font = df_font if df_font else bold
    c = ws.cell(row=row, column=6, value=totals["g_pct_hrs"]); c.font = bold; c.number_format = "0.00%"
    c = ws.cell(row=row, column=7, value=totals["total_drill"]); c.font = bold; c.number_format = COMMA_FMT
    c = ws.cell(row=row, column=8, value=int(round(grand_ftg_diff))); c.number_format = COMMA_FMT
    df_font = _diff_font(grand_ftg_diff, bold=True)
    c.font = df_font if df_font else bold
    c = ws.cell(row=row, column=9, value=totals["g_pct_drill"]); c.font = bold; c.number_format = "0.00%"
    ws.cell(row=row, column=10, value=totals["incidents"]).font = bold
    ws.cell(row=row, column=11, value=kpi_data.get("grand_top_operator", "")).font = bold
    ws.row_dimensions[row].height = DATA_ROW_HEIGHT

    # Color-scale gradients (per-hole-size rows only — excludes Grand Total)
    if data_last_row >= data_first_row:
        # More is better: low=red, mid=yellow, high=green  (Runs, Hrs, Drill)
        more_is_better = ColorScaleRule(
            start_type="min", start_color="F8696B",   # red (low)
            mid_type="percentile", mid_value=50, mid_color="FFEB84",  # yellow
            end_type="max", end_color="63BE7B",       # green (high)
        )
        # More is worse: low=green, mid=yellow, high=red  (Incidents)
        more_is_worse = ColorScaleRule(
            start_type="min", start_color="63BE7B",   # green (low)
            mid_type="percentile", mid_value=50, mid_color="FFEB84",  # yellow
            end_type="max", end_color="F8696B",       # red (high)
        )
        for col in (3, 4, 7):  # Total Runs, Total Hrs, Total Drill
            letter = get_column_letter(col)
            ws.conditional_formatting.add(
                f"{letter}{data_first_row}:{letter}{data_last_row}", more_is_better,
            )
        # Total Incidents
        letter = get_column_letter(10)
        ws.conditional_formatting.add(
            f"{letter}{data_first_row}:{letter}{data_last_row}", more_is_worse,
        )

    # Black outline around the entire summary table (incl. header and grand total)
    _apply_block_border(ws, start_row, row, 1, len(SUMMARY_HEADERS),
                        BLACK_BORDER, style="medium")

    return row + 1


def _write_detailed_table(ws, kpi_data, start_row, week_num):
    """Write the detailed (per motor-type) table starting at start_row.
    Longest Run rows are NOT included here — they go into the Longest Run Table.
    Returns the row immediately after."""
    bold_white = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")

    # Header row
    for c, h in enumerate(DETAILED_HEADERS, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.font = bold_white
        cell.fill = header_fill
        cell.alignment = CENTER_WRAP
    ws.row_dimensions[start_row].height = HEADER_ROW_HEIGHT

    row = start_row + 1
    flat_data = []          # (ws_row, row_dict) for global max/min lookup
    block_ranges = []       # list of (top_row, bottom_row) for black borders

    for block in kpi_data["blocks"]:
        block_top = row

        # Per-group data rows
        block_data_rows = []
        for r in block["rows"]:
            mkey = _normalize_motor_key(r.get("motor_type", ""))
            row_fill = PatternFill("solid", fgColor=MOTOR_FILL[mkey]) if mkey else None
            if row_fill:
                for c in range(1, DETAILED_LAST_COL + 1):
                    ws.cell(row=row, column=c).fill = row_fill

            ws.cell(row=row, column=1, value=week_num)
            ws.cell(row=row, column=2, value=block["hole_size"])
            ws.cell(row=row, column=3, value=r["motor_type"])
            ws.cell(row=row, column=4, value=r["job_type"])
            ws.cell(row=row, column=5, value=r["series_20"])
            ws.cell(row=row, column=6, value=r["runs"])
            ws.cell(row=row, column=7, value=r["total_hrs"])
            ws.cell(row=row, column=8, value=r["w_pct_hrs"]).number_format = "0.00%"
            ws.cell(row=row, column=9, value=r["g_pct_hrs"]).number_format = "0.00%"
            c = ws.cell(row=row, column=10, value=r["total_drill"]); c.number_format = COMMA_FMT
            ws.cell(row=row, column=11, value=r["w_pct_drill"]).number_format = "0.00%"
            ws.cell(row=row, column=12, value=r["g_pct_drill"]).number_format = "0.00%"
            ws.cell(row=row, column=13, value=r["drilling_hrs"])
            ws.cell(row=row, column=14, value=r["avg_rop"])
            ws.cell(row=row, column=15, value=r["avg_slide_pct"]).number_format = "0.00%"
            c = ws.cell(row=row, column=16, value=r["avg_run_length"]); c.number_format = COMMA_FMT
            ws.cell(row=row, column=17, value=r["my_avg"])
            ws.cell(row=row, column=18, value=r["incident_count"] if r["incident_count"] else None)
            ws.cell(row=row, column=19, value=r.get("top_operator", ""))

            ws.row_dimensions[row].height = DATA_ROW_HEIGHT
            block_data_rows.append((row, r))
            flat_data.append((row, r))
            row += 1

        # Per-hole-size winner: dark-fill the Total Drill cell using motor-type's dark shade
        if block_data_rows:
            best_idx = max(range(len(block_data_rows)),
                           key=lambda i: block_data_rows[i][1]["total_drill"])
            best_ws_row, best_row_data = block_data_rows[best_idx]
            mkey = _normalize_motor_key(best_row_data.get("motor_type", ""))
            dark = MOTOR_FILL_DARK.get(mkey)
            if dark:
                cell = ws.cell(row=best_ws_row, column=COL_TOTAL_DRILL)
                cell.fill = PatternFill("solid", fgColor=dark)
                cell.font = Font(bold=True, color="FFFFFF")

        block_bottom = row - 1
        if block_bottom >= block_top:
            block_ranges.append((block_top, block_bottom))

    # Center alignment on detailed table cells
    for r in range(start_row, row):
        for c in range(1, DETAILED_LAST_COL + 1):
            ws.cell(row=r, column=c).alignment = (
                CENTER_WRAP if r == start_row else CENTER
            )

    # Global highest/lowest borders across all data rows
    if flat_data:
        max_drill = max(flat_data, key=lambda x: x[1]["total_drill"])
        min_drill = min(flat_data, key=lambda x: x[1]["total_drill"])
        max_hrs = max(flat_data, key=lambda x: x[1]["total_hrs"])
        min_hrs = min(flat_data, key=lambda x: x[1]["total_hrs"])
        max_runlen = max(flat_data, key=lambda x: x[1]["avg_run_length"])
        min_runlen = min(flat_data, key=lambda x: x[1]["avg_run_length"])

        _apply_border_span(ws, max_drill[0], COL_HOLE_SIZE, COL_SERIES_20, GREEN_BORDER)
        _apply_border_single(ws, max_drill[0], COL_TOTAL_DRILL, GREEN_BORDER)
        _apply_border_single(ws, max_drill[0], COL_PCT_G_DRILL, GREEN_BORDER)
        _apply_border_span(ws, min_drill[0], COL_HOLE_SIZE, COL_SERIES_20, RED_BORDER)
        _apply_border_single(ws, min_drill[0], COL_TOTAL_DRILL, RED_BORDER)
        _apply_border_single(ws, min_drill[0], COL_PCT_G_DRILL, RED_BORDER)

        _apply_border_single(ws, max_hrs[0], COL_TOTAL_HRS, GREEN_BORDER)
        _apply_border_single(ws, max_hrs[0], COL_PCT_G_HRS, GREEN_BORDER)
        _apply_border_single(ws, min_hrs[0], COL_TOTAL_HRS, RED_BORDER)
        _apply_border_single(ws, min_hrs[0], COL_PCT_G_HRS, RED_BORDER)

        _apply_border_single(ws, max_runlen[0], COL_AVG_RUNLEN, GREEN_BORDER)
        _apply_border_single(ws, min_runlen[0], COL_AVG_RUNLEN, RED_BORDER)

    # Black border around each hole-size block
    for top, bottom in block_ranges:
        _apply_block_border(ws, top, bottom, 1, DETAILED_LAST_COL,
                            BLACK_BORDER, style="medium")

    return row


def _write_longest_run_table(ws, kpi_data, start_row, week_num):
    """Write the Longest Run table — one row per hole size, packed into cols A..M.
    Returns the row immediately after."""
    bold_white = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")

    # Header row
    for c, label in enumerate(LONGEST_RUN_HEADERS, start=1):
        cell = ws.cell(row=start_row, column=c, value=label)
        cell.font = bold_white
        cell.fill = header_fill
        cell.alignment = CENTER_WRAP
    ws.row_dimensions[start_row].height = HEADER_ROW_HEIGHT

    row = start_row + 1

    # Collect entries and sort by Total Drill descending
    lr_entries = []
    for block in kpi_data["blocks"]:
        lr = block.get("longest_run")
        if lr:
            lr_entries.append((block["hole_size"], lr))
    lr_entries.sort(key=lambda x: -x[1].get("total_drill", 0))

    for hs, lr in lr_entries:
        mkey = _normalize_motor_key(lr.get("motor_type", ""))
        row_fill_color = MOTOR_FILL.get(mkey) if mkey else None
        if row_fill_color:
            for c in range(1, LONGEST_LAST_COL + 1):
                ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=row_fill_color)

        ws.cell(row=row, column=LR_COL_WEEK, value=week_num)
        ws.cell(row=row, column=LR_COL_HOLE_SIZE, value=hs)
        ws.cell(row=row, column=LR_COL_MOTOR_TYPE, value=lr["motor_type"])
        ws.cell(row=row, column=LR_COL_JOB_TYPE, value=lr["job_type"])
        ws.cell(row=row, column=LR_COL_SERIES_20, value=lr["series_20"])
        ws.cell(row=row, column=LR_COL_TOTAL_HRS, value=lr["total_hrs"])
        c = ws.cell(row=row, column=LR_COL_TOTAL_DRILL, value=lr["total_drill"])
        c.number_format = COMMA_FMT
        ws.cell(row=row, column=LR_COL_DRILLING_HRS, value=lr["drilling_hrs"])
        ws.cell(row=row, column=LR_COL_AVG_ROP, value=lr["avg_rop"])
        ws.cell(row=row, column=LR_COL_AVG_SLIDE, value=lr["avg_slide_pct"]).number_format = "0.00%"
        ws.cell(row=row, column=LR_COL_MY_AVG, value=lr["my_avg"])
        ws.cell(row=row, column=LR_COL_INCIDENT,
                value=lr["incident_count"] if lr["incident_count"] else None)
        ws.cell(row=row, column=LR_COL_OPERATOR, value=lr.get("operator", ""))
        ws.cell(row=row, column=LR_COL_JOB_NUM, value=lr.get("job_num", ""))
        ws.cell(row=row, column=LR_COL_PHASE, value=lr.get("phase", ""))
        ws.cell(row=row, column=LR_COL_BEND, value=lr.get("bend", ""))

        ws.row_dimensions[row].height = DATA_ROW_HEIGHT

        for c in range(1, LONGEST_LAST_COL + 1):
            ws.cell(row=row, column=c).alignment = CENTER

        row += 1

    table_bottom = row - 1

    # Black border around the whole table (header + data rows)
    if table_bottom >= start_row:
        _apply_block_border(ws, start_row, table_bottom, 1, LONGEST_LAST_COL,
                            BLACK_BORDER, style="medium")

    return row


def write_kpi_excel(kpi_data, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly KPI"

    week_num = _week_num(kpi_data.get("week", ""))

    # 1) Summary table at the top
    next_row = _write_summary_table(ws, kpi_data, start_row=1, week_num=week_num)

    # 2) Two-row gap
    next_row += 2

    # 3) Detailed table
    next_row = _write_detailed_table(ws, kpi_data, start_row=next_row, week_num=week_num)

    # 4) Two-row gap
    next_row += 2

    # 5) Longest Run table
    _write_longest_run_table(ws, kpi_data, start_row=next_row, week_num=week_num)

    # Column widths
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    wb.save(path)
    return path
